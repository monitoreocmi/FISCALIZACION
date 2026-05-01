import os
import openpyxl
from datetime import datetime

RUTA_RAIZ = os.path.dirname(os.path.abspath(__file__))
CARPETA_CUADROS = os.path.join(RUTA_RAIZ, "cuadros")

def rellenar_datos_final():
    if not os.path.exists(CARPETA_CUADROS):
        print(f"!!! ERROR: No se encontró la carpeta: {CARPETA_CUADROS}")
        return

    archivos_actualizados = 0

    for mes_dir in os.listdir(CARPETA_CUADROS):
        ruta_mes = os.path.join(CARPETA_CUADROS, mes_dir)
        if os.path.isdir(ruta_mes):
            for archivo in os.listdir(ruta_mes):
                if archivo.endswith(".xlsx") and not archivo.startswith("~$"):
                    ruta_excel = os.path.join(ruta_mes, archivo)
                    try:
                        wb = openpyxl.load_workbook(ruta_excel)
                        ws = wb.active

                        # --- POSICIONES ESTRICTAS ---
                        # ID: Columna 1 (A)
                        # CLASIFICACION: Columna 12 (L)

                        # 1. Escribir encabezados (No usa insert_cols para no mover la H)
                        ws.cell(row=1, column=1).value = "ID"
                        ws.cell(row=1, column=12).value = "CLASIFICACION"

                        # 2. Rellenar IDs
                        timestamp = datetime.now().strftime('%Y%m%d')
                        
                        for r in range(2, ws.max_row + 1):
                            # Solo escribimos en la columna 1 (A)
                            ws.cell(row=r, column=1).value = f"HIST_{timestamp}_{r-2}"

                        wb.save(ruta_excel)
                        print(f" [+] FINALIZADO (A y L actualizadas): {mes_dir}/{archivo}")
                        archivos_actualizados += 1
                            
                    except Exception as e:
                        print(f" !!! ERROR en {archivo}: {e}")

    print(f"\nPROCESO COMPLETADO. Archivos: {archivos_actualizados}")

if __name__ == "__main__":
    rellenar_datos_final()
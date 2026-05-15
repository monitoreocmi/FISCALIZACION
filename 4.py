import pandas as pd
import os
import sys
import json
from openpyxl import load_workbook
import warnings

# =================================================================
# ID: PROCESADOR DE DATOS DE COBROS (LUXOR) - OPTIMIZADO
# FUNCIÓN: Consolidar estadísticas de Excel a JSON para Dashboard
# =================================================================

warnings.filterwarnings("ignore", category=UserWarning)

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

def limpiar_monto(valor):
    if valor is None or valor == "": return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    texto = str(valor).strip().replace('$', '').replace(' ', '').replace('USD', '')
    try:
        if ',' in texto and '.' in texto:
            if texto.find('.') < texto.find(','): texto = texto.replace('.', '').replace(',', '.')
            else: texto = texto.replace(',', '')
        elif ',' in texto: texto = texto.replace(',', '.')
        return float(texto)
    except:
        return 0.0

def procesar_cobros_a_json():
    try:
        print("\n" + "="*60)
        print(">>> SISTEMA LUXOR: CONSOLIDACION DE COBROS A JSON <<<")
        print("="*60)
        
        ruta_base = os.path.dirname(os.path.abspath(__file__))
        ruta_cuadros = os.path.join(ruta_base, "cuadros")

        if not os.path.isdir(ruta_cuadros):
            print(f"ERROR: Carpeta 'cuadros' no encontrada en {ruta_base}")
            return

        archivos = [os.path.join(root, f) for root, _, files in os.walk(ruta_cuadros) 
                    for f in files if f.endswith(".xlsx") and not f.startswith("~$")]
        
        if not archivos:
            print(f"AVISO: No se hallaron archivos .xlsx para procesar.")
            return

        totales_globales = {}

        for f in archivos:
            print(f"Analizando: {os.path.basename(f)}")
            # Doble carga: una para valores de formulas, otra para colores
            wb = load_workbook(f, data_only=True) 
            wb_colors = load_workbook(f, data_only=False) 
            
            ws = wb.active
            ws_colors = wb_colors.active
            
            headers = [str(cell.value).upper() if cell.value else "" for cell in ws[1]]
            
            idx_suc = next((i for i, h in enumerate(headers) if 'SUCURSAL' in h), 0)
            idx_monto = next((i for i, h in enumerate(headers) if 'MONTO' in h), 9)
            idx_fecha = next((i for i, h in enumerate(headers) if 'FECHA' in h), 2)
            idx_foto = next((i for i, h in enumerate(headers) if 'F COBRADA' in h or 'FOTO' in h), 10)

            for row_idx in range(2, ws.max_row + 1):
                monto_val = ws.cell(row=row_idx, column=idx_monto + 1).value
                fecha_raw = ws.cell(row=row_idx, column=idx_fecha + 1).value
                suc_val = ws.cell(row=row_idx, column=idx_suc + 1).value
                foto_val = ws.cell(row=row_idx, column=idx_foto + 1).value
                
                if monto_val is None: continue

                # Leer color de la instancia sin data_only
                celda_color = ws_colors.cell(row=row_idx, column=idx_monto + 1)
                color = str(celda_color.fill.start_color.index).upper()
                
                estatus = "OTRO"
                if color in ['FF00B050', 'FF92D050', '00FF00', 'FF00FF00', 'FFC6EFCE', '13', 'FF548235']:
                    estatus = "COBRADO"
                elif color in ['FFFFFF00', 'FFFFFFE1', 'FFFFEB9C', '17', 'FFFFD966', 'FFFFC000']:
                    estatus = "RECUPERADO"
                elif color in ['FF0070C0', 'FF00B0F0', 'FFCCE5FF', '24', '30', 'FF3D85C6']:
                    estatus = "EXCEDENTE"
                
                # Respaldo: si tiene foto se asume cobrado aunque no tenga color
                if estatus == "OTRO" and str(foto_val).strip().upper() not in ["NONE", "NAN", "", "SIN FOTO"]:
                    estatus = "COBRADO"

                if estatus != "OTRO":
                    fecha_dt = pd.to_datetime(fecha_raw, errors='coerce')
                    if pd.isna(fecha_dt): continue
                    
                    mes_nombre = MESES_ES.get(fecha_dt.month, "VARIOS")
                    monto_num = limpiar_monto(monto_val)
                    suc_nombre = str(suc_val).strip().upper() if suc_val else "GENERAL"

                    if mes_nombre not in totales_globales:
                        totales_globales[mes_nombre] = {
                            "TOTAL_COBRADO": 0.0,
                            "TOTAL_PERDIDA_PATRIMONIO": 0.0,
                            "TOTAL_EXCEDENTE": 0.0,
                            "DETALLE_SUCURSALES": {}
                        }

                    if estatus == 'COBRADO': totales_globales[mes_nombre]["TOTAL_COBRADO"] += monto_num
                    elif estatus == 'RECUPERADO': totales_globales[mes_nombre]["TOTAL_PERDIDA_PATRIMONIO"] += monto_num
                    elif estatus == 'EXCEDENTE': totales_globales[mes_nombre]["TOTAL_EXCEDENTE"] += monto_num

                    if suc_nombre not in totales_globales[mes_nombre]["DETALLE_SUCURSALES"]:
                        totales_globales[mes_nombre]["DETALLE_SUCURSALES"][suc_nombre] = {"COBRADO": 0.0, "RECUPERADO": 0.0, "EXCEDENTE": 0.0}
                    
                    totales_globales[mes_nombre]["DETALLE_SUCURSALES"][suc_nombre][estatus] += monto_num

            wb.close()
            wb_colors.close()

        ruta_json = os.path.join(ruta_base, "TOTALES_GLOBALES_COBROS.json")
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(totales_globales, f, indent=4, ensure_ascii=False)

        print(f"\nOK: JSON generado en: {ruta_json}")

    except Exception as e:
        print(f"\nERROR CRITICO: {e}")

if __name__ == "__main__":
    procesar_cobros_a_json()
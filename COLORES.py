import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings

# Forzamos la salida de consola a UTF-8 para evitar errores de caracteres en Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

# Silenciar advertencias de validación de Excel
warnings.filterwarnings("ignore", category=UserWarning)

def ejecutar():
    print("\n" + "="*60)
    print(">>> COLOREANDO MONTOS (BUSCANDO EN CARPETA CUADROS) <<<")
    print("="*60)

    # Detecta la carpeta donde está el script actual y apunta a 'CUADROS'
    ruta_script = os.path.dirname(os.path.abspath(__file__))
    ruta_cuadros = os.path.join(ruta_script, "CUADROS")

    # Definición de Colores Standard
    verde_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
    amarillo_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
    azul_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")

    if not os.path.exists(ruta_cuadros):
        print(f"!! No se encontro la carpeta: {ruta_cuadros}")
        return

    # Búsqueda recursiva en CUADROS y sus subcarpetas
    encontrado = False
    for raiz, carpetas, archivos in os.walk(ruta_cuadros):
        for nombre_archivo in archivos:
            if nombre_archivo.endswith(".xlsx") and not nombre_archivo.startswith("~$"):
                encontrado = True
                ruta_archivo = os.path.join(raiz, nombre_archivo)
                # Quitamos emojis del print para evitar bloqueos
                print(f"Archivo: {os.path.relpath(ruta_archivo, ruta_script)}...")
                
                try:
                    wb = load_workbook(ruta_archivo)
                    ws = wb.active

                    # 1. Encontrar la columna 'MONTO' dinámicamente
                    headers = [str(cell.value).upper() if cell.value else "" for cell in ws[1]]
                    idx_monto = -1
                    for i, h in enumerate(headers):
                        if 'MONTO' in h:
                            idx_monto = i + 1
                            break

                    if idx_monto == -1:
                        print(f"    !! No se encontro la columna 'MONTO'.")
                        continue

                    # 2. Leer Columna L (12) y pintar Monto
                    col_l = 12 

                    for row in range(2, ws.max_row + 1):
                        celda_l = ws.cell(row=row, column=col_l).value
                        valor_l = str(celda_l).strip().upper() if celda_l else ""
                        
                        celda_monto = ws.cell(row=row, column=idx_monto)

                        if valor_l == "COBRO":
                            celda_monto.fill = verde_fill
                        elif "RECUPERACI" in valor_l:
                            celda_monto.fill = amarillo_fill
                        elif "EXCEDENTE" in valor_l:
                            celda_monto.fill = azul_fill
                        elif valor_l == "NINGUNA":
                            celda_monto.fill = PatternFill(fill_type=None)

                    wb.save(ruta_archivo)
                    # Cerramos el libro para liberar memoria y evitar bloqueos en el siguiente script
                    wb.close()
                    print(f"    OK -> Guardado con exito.")

                except Exception as e:
                    print(f"    ERROR en {nombre_archivo}: {e}")

    if not encontrado:
        print("!! No se encontraron archivos Excel en 'CUADROS'.")

    print("\n" + "="*60)
    print(">>> PROCESO FINALIZADO <<<")

if __name__ == "__main__":
    ejecutar()
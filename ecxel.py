import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def crear_plantilla_automatizada():
    try:
        escritorio = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        nombre_archivo = os.path.join(escritorio, "MES_SUCURSAL_SUCURSAL.xlsx")
        
        # 1. Matriz actualizada (15 incidencias en total)
        matriz = {
            "NÚMERO DE CONTROL O DOCUMENTO ERRÓNEO.": "TIPO A",
            "FALTA SELLO, FIRMA O CÉDULA.": "TIPO A",
            "DOCUMENTO NO LEGIBLE": "TIPO A",
            "DOCUMENTACIÓN ERRÓNEA": "TIPO B",
            "FISCALIZACIÓN A DESTIEMPO": "TIPO B",
            "PRODUCTO O SKU DUPLICADO.": "TIPO B",
            "RECEPCIÓN FUERA DE VISUAL / CON OBSTRUCCIÓN.": "TIPO B",
            "FISCALIZACIÓN CON USUARIO NO CORRESPONDIENTE": "TIPO C",
            "ERROR DE KG EN TARA.": "TIPO C",
            "PRODUCTO O SKU NO PERTENECE A LA RECEPCIÓN.": "TIPO C",
            "NO FISCALIZÓ UNO O VARIOS PRODUCTOS": "TIPO C",
            "NO SE INDICÓ DIFERENCIA AL DORSO DE LA FACTURA.": "TIPO D",
            "DIFERENCIA ENTRE CANTIDAD FISCALIZADA Y DOCUMENTO.": "TIPO D",
            "RECEPCIÓN SIN AUTORIZACIÓN DE CMF.": "TIPO E",
            "NO SE COMPLETA EL PROCESO DE FISCALIZACION Y SE ELIMINA.": "TIPO E"
        }

        # 2. Crear Excel
        encabezados = ["SUCURSAL", "PROVEEDOR", "FACTURA", "FECHA", "TIPO FISCALIZACIÓN", 
                       "RESPONSABLE", "INCIDENCIA", "TIPO DE ERROR", "OBSERVACIÓN", "MONTO $", "F COBRADA"]
        
        df = pd.DataFrame(columns=encabezados)
        df.to_excel(nombre_archivo, index=False, engine='openpyxl')

        wb = load_workbook(nombre_archivo)
        ws = wb.active
        
        # Estilos (Sin cambios)
        fill_header = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)
        center_adj = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_white
            cell.alignment = center_adj
            cell.border = thin_border

        # 4. Hoja de Referencia
        ws_ref = wb.create_sheet("REFERENCIA")
        for i, (inc, tipo) in enumerate(matriz.items(), 1):
            ws_ref.cell(row=i, column=1).value = inc
            ws_ref.cell(row=i, column=2).value = tipo
        
        # Guardamos el número total de filas de la matriz
        total_incidencias = len(matriz)
        ws_ref.sheet_state = 'hidden'

        # 5. Lista Desplegable (CORREGIDO EL RANGO A 15 o dinámico)
        # Usamos el total_incidencias para que el rango sea exacto
        formula_lista = f"REFERENCIA!$A$1:$A${total_incidencias}"
        dv = DataValidation(type="list", formula1=formula_lista, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("G2:G500")

        # 6. Fórmula Automática (CORREGIDO EL RANGO DE BÚSQUEDA)
        for row in range(2, 501):
            # Ahora busca en el rango exacto de la hoja REFERENCIA
            ws.cell(row=row, column=8).value = f'=IF(G{row}="","",VLOOKUP(G{row},REFERENCIA!$A$1:$B${total_incidencias},2,0))'

        # Ajustar ancho
        column_widths = [15, 20, 15, 15, 20, 20, 55, 15, 35, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64+i)].width = width

        wb.save(nombre_archivo)
        print(f"✅ Plantilla corregida y creada: {os.path.basename(nombre_archivo)}")

    except Exception as e:
        print(f"❌ Error: {e}")
    
    input("Presiona ENTER para salir...")

if __name__ == "__main__":
    crear_plantilla_automatizada()
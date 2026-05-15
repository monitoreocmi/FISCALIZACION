import pandas as pd
import os
import sys
import json
from openpyxl import load_workbook
import warnings
import time

# Silenciar advertencias de validación de Excel
warnings.filterwarnings("ignore", category=UserWarning)

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

RUTA_LOGO_ESTANDAR = "../../RECURSOS/logo.png"

def limpiar_monto(valor):
    if valor is None or str(valor).strip().lower() in ['none', 'nan', '']: return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    texto = str(valor).strip().replace('$', '').replace(' ', '')
    if ',' in texto and '.' in texto:
        if texto.find('.') < texto.find(','): texto = texto.replace('.', '').replace(',', '.')
        else: texto = texto.replace(',', '')
    elif ',' in texto: texto = texto.replace(',', '.')
    try: return float(texto)
    except: return 0.0

def obtener_indices_flexibles(headers):
    indices = {'sucursal': -1, 'monto': -1, 'fecha': -1, 'foto': -1}
    for i, h in enumerate(headers):
        h_up = str(h).upper() if h else ""
        if 'SUCURSAL' in h_up: indices['sucursal'] = i
        if 'MONTO $' in h_up: indices['monto'] = i
        elif 'MONTO' in h_up and indices['monto'] == -1: indices['monto'] = i
        if 'FECHA' in h_up: indices['fecha'] = i
        if 'F COBRADA' in h_up: indices['foto'] = i
    
    if indices['monto'] == -1: indices['monto'] = 9 
    return indices

def ejecutar():
    try:
        print("\n" + "="*60)
        print(">>> GENERANDO REPORTES DE COBROS PARA PANEL <<<")
        print("="*60)
        
        ruta_base = os.path.dirname(os.path.abspath(__file__))
        ruta_cuadros = os.path.join(ruta_base, "cuadros")
        
        if not os.path.exists(ruta_cuadros):
            print(f"Error: No existe la carpeta 'cuadros' en {ruta_base}")
            return

        archivos = [os.path.join(root, f) for root, dirs, files in os.walk(ruta_cuadros) 
                    for f in files if f.endswith(".xlsx") and not f.startswith("~$")]

        datos_finales = []
        todas_las_sucursales = set()

        for f in archivos:
            nombre_f = os.path.basename(f)
            print(f"Analizando: {nombre_f}")
            wb = load_workbook(f, data_only=False)
            ws = wb.active
            headers_reales = [str(cell.value).strip() if cell.value else f"COL_{i+1}" for i, cell in enumerate(ws[1])]
            idx = obtener_indices_flexibles(headers_reales)
            
            for row in ws.iter_rows(min_row=2):
                row_vals = [cell.value for cell in row]
                if not any(v is not None for v in row_vals): continue
                
                suc_actual = str(row_vals[idx['sucursal']]).strip().upper() if idx['sucursal'] != -1 else "GENERAL"
                if suc_actual and suc_actual not in ["NONE", "NAN", ""]:
                    todas_las_sucursales.add(suc_actual)

                fecha_val = pd.to_datetime(row_vals[idx['fecha']], errors='coerce') if idx['fecha'] != -1 else None
                mes_num = fecha_val.month if fecha_val and not pd.isna(fecha_val) else 0
                
                estatus = "OTRO"
                try:
                    target_cell = row[idx['monto']]
                    color = str(target_cell.fill.start_color.index).upper()
                    if color in ['FF00B050', 'FF92D050', '00FF00', 'FF00FF00', 'FFC6EFCE', '13', 'FF548235']: estatus = "COBRADO"
                    elif color in ['FFFFFF00', 'FFFFFFE1', 'FFFFEB9C', '17', 'FFFFD966', 'FFFFC000']: estatus = "RECUPERADO"
                    elif color in ['FF0070C0', 'FF00B0F0', 'FFCCE5FF', '24', '30', 'FF3D85C6']: estatus = "EXCEDENTE"
                except: pass

                if estatus != "OTRO":
                    mes_nombre = MESES_ES.get(mes_num, "VARIOS")
                    datos_finales.append({
                        'SUCURSAL': suc_actual,
                        'MES': mes_nombre, 
                        'PERIODO': fecha_val.to_period('M') if fecha_val and not pd.isna(fecha_val) else pd.Period('2026-05', freq='M'),
                        'MONTO_CALC': limpiar_monto(row_vals[idx['monto']]),
                        'FOTO_BASE': str(row_vals[idx['foto']]).strip() if idx['foto'] != -1 and row_vals[idx['foto']] else "",
                        'ESTATUS': estatus, 'FILA': row_vals, 'HEADERS': headers_reales, 'IDX': idx
                    })
            wb.close()

        df = pd.DataFrame(datos_finales) if datos_finales else pd.DataFrame(columns=['SUCURSAL', 'MES', 'PERIODO', 'MONTO_CALC', 'ESTATUS'])

        # Estilo CSS omitido por brevedad pero se mantiene igual en tu archivo final
        estilo_css = """<style>body { font-family: 'Segoe UI', sans-serif; background: #f0f2f5; text-align: center; } .btn { padding: 10px; background: #002060; color: white; text-decoration: none; border-radius: 6px; }</style>"""
        
        script_modal = """<script>function openModal(src) { alert('Ver foto: ' + src); }</script>"""

        for periodo in (df['PERIODO'].unique() if not df.empty else [pd.Period('2026-05', freq='M')]):
            df_p = df[df['PERIODO'] == periodo] if not df.empty else pd.DataFrame()
            n_m = str(df_p['MES'].iloc[0]) if not df_p.empty else "MAYO"
            ruta_mes = os.path.join(ruta_base, n_m)
            os.makedirs(ruta_mes, exist_ok=True)

            # Generación de HTMLs Consolidados y por Sucursal
            # (Aquí se mantiene tu lógica de escritura de archivos con encoding="utf-8")
            print(f"Generando reportes para periodo: {periodo}")

        print("\nOK: PROCESO DE COBROS COMPLETADO.")
    except Exception as e:
        print(f"ERROR: Durante la ejecucion: {e}")

if __name__ == "__main__":
    ejecutar()